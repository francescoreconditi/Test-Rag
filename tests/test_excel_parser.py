# ============================================
# FILE DI TEST/DEBUG - NON PER PRODUZIONE
# Creato da: Claude Code
# Data: 2025-09-08
# Scopo: Test per il parser Excel con metadati completi
# ============================================

"""Test per ExcelParser."""

import tempfile
from pathlib import Path
from datetime import datetime

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from src.application.parsers.excel_parser import (
    ExcelParser,
    CellMetadata,
    TableMetadata,
    SheetMetadata,
    WorkbookMetadata,
    ExtractedData
)


@pytest.fixture
def sample_excel_file():
    """Crea un file Excel di esempio per i test."""
    # Crea workbook
    wb = Workbook()
    
    # Sheet 1: Dati finanziari
    ws1 = wb.active
    ws1.title = "Bilancio"
    
    # Headers
    ws1['A1'] = 'Voce'
    ws1['B1'] = '2023'
    ws1['C1'] = '2024'
    ws1['D1'] = 'Variazione %'
    
    # Dati
    ws1['A2'] = 'Ricavi'
    ws1['B2'] = 1000000
    ws1['C2'] = 1200000
    ws1['D2'] = '=((C2-B2)/B2)*100'
    
    ws1['A3'] = 'Costi'
    ws1['B3'] = 800000
    ws1['C3'] = 900000
    ws1['D3'] = '=((C3-B3)/B3)*100'
    
    ws1['A4'] = 'EBITDA'
    ws1['B4'] = '=B2-B3'
    ws1['C4'] = '=C2-C3'
    ws1['D4'] = '=((C4-B4)/B4)*100'
    
    # Totali
    ws1['A6'] = 'TOTALE'
    ws1['B6'] = '=SUM(B2:B3)'
    ws1['C6'] = '=SUM(C2:C3)'
    
    # Aggiungi commento
    from openpyxl.comments import Comment
    ws1['A2'].comment = Comment("Ricavi delle vendite e prestazioni", "Test Author")
    
    # Formattazione numeri
    for col in ['B', 'C']:
        for row in range(2, 7):
            cell = ws1[f'{col}{row}']
            if cell.value:
                cell.number_format = '#,##0'
    
    ws1['D2'].number_format = '0.00%'
    ws1['D3'].number_format = '0.00%'
    ws1['D4'].number_format = '0.00%'
    
    # Sheet 2: Dati HR
    ws2 = wb.create_sheet("HR")
    
    ws2['A1'] = 'Reparto'
    ws2['B1'] = 'FTE'
    ws2['C1'] = 'Costo Medio'
    
    ws2['A2'] = 'Vendite'
    ws2['B2'] = 50
    ws2['C2'] = 45000
    
    ws2['A3'] = 'Amministrazione'
    ws2['B3'] = 20
    ws2['C3'] = 40000
    
    ws2['A4'] = 'Produzione'
    ws2['B4'] = 100
    ws2['C4'] = 35000
    
    # Celle unite
    ws2.merge_cells('A6:C6')
    ws2['A6'] = 'Dati al 31/12/2024'
    
    # Salva in file temporaneo
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        return Path(tmp.name)


def test_excel_parser_initialization():
    """Test inizializzazione parser."""
    parser = ExcelParser()
    assert parser.workbook is None
    assert parser.file_path is None


def test_parse_excel_file(sample_excel_file):
    """Test parsing completo file Excel."""
    parser = ExcelParser()
    extracted_data = parser.parse(sample_excel_file)
    
    # Verifica struttura dati estratti
    assert isinstance(extracted_data, ExtractedData)
    assert isinstance(extracted_data.workbook_metadata, WorkbookMetadata)
    
    # Verifica metadati workbook
    assert extracted_data.workbook_metadata.file_path == str(sample_excel_file)
    assert extracted_data.workbook_metadata.sheets_count == 2
    assert len(extracted_data.workbook_metadata.sheets) == 2
    
    # Verifica dataframes
    assert 'Bilancio' in extracted_data.data_frames
    assert 'HR' in extracted_data.data_frames
    
    df_bilancio = extracted_data.data_frames['Bilancio']
    assert df_bilancio.shape[0] >= 6  # Almeno 6 righe
    
    # Verifica cell metadata
    assert 'Bilancio' in extracted_data.cell_metadata
    bilancio_cells = extracted_data.cell_metadata['Bilancio']
    assert len(bilancio_cells) > 0
    
    # Verifica formule estratte
    assert 'Bilancio' in extracted_data.formulas
    formulas = extracted_data.formulas['Bilancio']
    assert len(formulas) > 0
    
    # Verifica che le formule siano state rilevate
    formula_cells = [f[0] for f in formulas]
    assert 'D2' in formula_cells or 'B4' in formula_cells
    
    # Verifica commenti
    assert 'Bilancio' in extracted_data.comments
    comments = extracted_data.comments['Bilancio']
    assert len(comments) > 0
    assert any('Ricavi' in comment[1] for comment in comments)


def test_workbook_metadata_extraction(sample_excel_file):
    """Test estrazione metadati workbook."""
    parser = ExcelParser()
    extracted_data = parser.parse(sample_excel_file)
    
    metadata = extracted_data.workbook_metadata
    
    # Verifica campi base
    assert metadata.file_path == str(sample_excel_file)
    assert metadata.file_hash is not None
    assert len(metadata.file_hash) == 64  # SHA256 hash
    assert metadata.file_size > 0
    assert metadata.sheets_count == 2
    
    # Verifica sheets metadata
    assert len(metadata.sheets) == 2
    
    bilancio_sheet = metadata.sheets[0]
    assert bilancio_sheet.name == 'Bilancio'
    assert bilancio_sheet.index == 0
    assert bilancio_sheet.visible is True
    assert bilancio_sheet.max_row >= 6
    assert bilancio_sheet.max_column >= 4
    
    hr_sheet = metadata.sheets[1]
    assert hr_sheet.name == 'HR'
    assert hr_sheet.index == 1
    assert len(hr_sheet.merged_cells_ranges) > 0  # Ha celle unite


def test_cell_metadata_extraction(sample_excel_file):
    """Test estrazione metadati celle."""
    parser = ExcelParser()
    extracted_data = parser.parse(sample_excel_file)
    
    bilancio_cells = extracted_data.cell_metadata['Bilancio']
    
    # Trova cella specifica
    ricavi_cell = next(
        (c for c in bilancio_cells if c.cell_reference == 'A2'),
        None
    )
    
    assert ricavi_cell is not None
    assert ricavi_cell.value == 'Ricavi'
    assert ricavi_cell.sheet_name == 'Bilancio'
    assert ricavi_cell.row == 2
    assert ricavi_cell.column == 1
    assert ricavi_cell.comment == "Ricavi delle vendite e prestazioni"
    
    # Verifica source reference
    source_ref = ricavi_cell.to_source_ref(str(sample_excel_file))
    assert f"{sample_excel_file}" in source_ref
    assert "sheet:Bilancio" in source_ref
    assert "cell:A2" in source_ref


def test_table_detection(sample_excel_file):
    """Test rilevamento tabelle."""
    parser = ExcelParser()
    extracted_data = parser.parse(sample_excel_file)
    
    # Dovrebbe rilevare almeno una tabella per sheet
    assert len(extracted_data.tables) >= 1
    
    # Verifica prima tabella
    if extracted_data.tables:
        table = extracted_data.tables[0]
        assert isinstance(table, TableMetadata)
        assert table.sheet_name in ['Bilancio', 'HR']
        assert len(table.headers) >= 3
        assert table.total_rows > 0
        assert table.total_columns >= 3
        
        # Verifica source reference
        source_ref = table.to_source_ref(str(sample_excel_file))
        assert f"{sample_excel_file}" in source_ref
        assert f"sheet:{table.sheet_name}" in source_ref


def test_formula_extraction(sample_excel_file):
    """Test estrazione formule."""
    parser = ExcelParser()
    extracted_data = parser.parse(sample_excel_file)
    
    # Verifica formule nel foglio Bilancio
    assert 'Bilancio' in extracted_data.formulas
    formulas = extracted_data.formulas['Bilancio']
    
    # Dovremmo avere diverse formule
    assert len(formulas) >= 5  # D2, D3, D4, B4, C4, B6, C6
    
    # Verifica struttura formula
    for cell_ref, formula in formulas:
        assert isinstance(cell_ref, str)
        assert isinstance(formula, str)
        assert formula.startswith('=')
    
    # Verifica formule specifiche
    formula_dict = dict(formulas)
    if 'B4' in formula_dict:
        assert 'B2-B3' in formula_dict['B4']


def test_comment_extraction(sample_excel_file):
    """Test estrazione commenti."""
    parser = ExcelParser()
    extracted_data = parser.parse(sample_excel_file)
    
    # Verifica commenti nel foglio Bilancio
    assert 'Bilancio' in extracted_data.comments
    comments = extracted_data.comments['Bilancio']
    
    assert len(comments) > 0
    
    # Verifica struttura commento
    for cell_ref, comment_text in comments:
        assert isinstance(cell_ref, str)
        assert isinstance(comment_text, str)
    
    # Verifica commento specifico
    comment_dict = dict(comments)
    assert 'A2' in comment_dict
    assert 'Ricavi' in comment_dict['A2']


def test_merged_cells_detection(sample_excel_file):
    """Test rilevamento celle unite."""
    parser = ExcelParser()
    extracted_data = parser.parse(sample_excel_file)
    
    # Verifica celle unite nel foglio HR
    hr_sheet = extracted_data.workbook_metadata.sheets[1]
    assert len(hr_sheet.merged_cells_ranges) > 0
    
    # Verifica nei metadati celle
    hr_cells = extracted_data.cell_metadata['HR']
    merged_cells = [c for c in hr_cells if c.is_merged]
    
    assert len(merged_cells) > 0
    
    # Verifica che il range sia specificato
    for cell in merged_cells:
        assert cell.merge_range is not None


def test_source_references_generation(sample_excel_file):
    """Test generazione source references."""
    parser = ExcelParser()
    data, source_refs = parser.extract_with_source_references(sample_excel_file)
    
    assert len(source_refs) > 0
    
    # Verifica struttura source references
    for ref in source_refs:
        assert ref.file_path == str(sample_excel_file)
        assert ref.page_number is None  # Excel non ha pagine
        assert ref.section is not None
        assert ref.metadata is not None
    
    # Verifica diversi tipi di source refs
    table_refs = [r for r in source_refs if 'Table' in r.subsection or 'Range' in r.subsection]
    cell_refs = [r for r in source_refs if 'Cell' in r.subsection]
    
    assert len(table_refs) > 0
    # Solo celle con formule o commenti generano source refs
    if len(cell_refs) > 0:
        assert any(r.metadata.get('formula') or r.metadata.get('comment') for r in cell_refs)


def test_raw_values_extraction(sample_excel_file):
    """Test estrazione valori raw."""
    parser = ExcelParser()
    extracted_data = parser.parse(sample_excel_file)
    
    # Verifica raw values per ogni sheet
    assert 'Bilancio' in extracted_data.raw_values
    assert 'HR' in extracted_data.raw_values
    
    bilancio_values = extracted_data.raw_values['Bilancio']
    
    # Verifica valori specifici
    assert 'A1' in bilancio_values
    assert bilancio_values['A1'] == 'Voce'
    
    assert 'A2' in bilancio_values
    assert bilancio_values['A2'] == 'Ricavi'
    
    assert 'B2' in bilancio_values
    assert bilancio_values['B2'] == 1000000


def test_number_format_extraction(sample_excel_file):
    """Test estrazione formato numeri."""
    parser = ExcelParser()
    extracted_data = parser.parse(sample_excel_file)
    
    bilancio_cells = extracted_data.cell_metadata['Bilancio']
    
    # Trova celle con formati numerici
    formatted_cells = [c for c in bilancio_cells if c.number_format]
    
    assert len(formatted_cells) > 0
    
    # Verifica formati percentuali
    percent_cells = [
        c for c in formatted_cells 
        if c.cell_reference in ['D2', 'D3', 'D4']
    ]
    
    for cell in percent_cells:
        if cell.number_format:  # Potrebbe essere None se non formattato
            assert '%' in cell.number_format


def test_error_handling_nonexistent_file():
    """Test gestione errore file non esistente."""
    parser = ExcelParser()
    
    with pytest.raises(FileNotFoundError):
        parser.parse("nonexistent_file.xlsx")


def test_empty_excel_file():
    """Test con file Excel vuoto."""
    wb = Workbook()
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = Path(tmp.name)
    
    parser = ExcelParser()
    extracted_data = parser.parse(tmp_path)
    
    # Dovrebbe comunque estrarre metadati base
    assert extracted_data.workbook_metadata is not None
    assert extracted_data.workbook_metadata.sheets_count >= 1  # Almeno il foglio di default
    
    # Clean up
    tmp_path.unlink()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])