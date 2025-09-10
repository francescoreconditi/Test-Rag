# ============================================
# FILE DI TEST/DEBUG - NON PER PRODUZIONE
# Creato da: Claude Code
# Data: 2025-09-08
# Scopo: Esempio di utilizzo dei nuovi componenti Excel parser e Raw blocks extractor
# ============================================

"""
Esempio di utilizzo dei nuovi componenti:
1. Excel Parser con metadati completi
2. Raw Blocks Extractor per documenti
"""

import asyncio
from pathlib import Path
import tempfile

from openpyxl import Workbook
from openpyxl.comments import Comment

# Import dei nuovi componenti
from src.application.parsers.excel_parser import ExcelParser
from src.application.services.enterprise_orchestrator import EnterpriseOrchestrator
from src.application.services.raw_blocks_extractor import BlockType, RawBlocksExtractor


def create_sample_excel():
    """Crea un file Excel di esempio con dati finanziari."""
    wb = Workbook()

    # Sheet 1: Bilancio
    ws = wb.active
    ws.title = "Bilancio 2024"

    # Headers
    ws['A1'] = 'Voce di Bilancio'
    ws['B1'] = 'Valore 2023'
    ws['C1'] = 'Valore 2024'
    ws['D1'] = 'Variazione %'

    # Dati finanziari
    data = [
        ['Ricavi delle vendite', 5000000, 5500000, '=((C2-B2)/B2)*100'],
        ['Costi della produzione', 3800000, 4100000, '=((C3-B3)/B3)*100'],
        ['EBITDA', '=B2-B3', '=C2-C3', '=((C4-B4)/B4)*100'],
        ['Ammortamenti', 200000, 220000, '=((C5-B5)/B5)*100'],
        ['EBIT', '=B4-B5', '=C4-C5', '=((C6-B6)/B6)*100'],
        ['Oneri finanziari', 50000, 45000, '=((C7-B7)/B7)*100'],
        ['Utile ante imposte', '=B6-B7', '=C6-C7', '=((C8-B8)/B8)*100'],
        ['Imposte', 300000, 350000, '=((C9-B9)/B9)*100'],
        ['Utile netto', '=B8-B9', '=C8-C9', '=((C10-B10)/B10)*100']
    ]

    for i, row in enumerate(data, start=2):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

    # Aggiungi commenti
    ws['A2'].comment = Comment("Include vendite Italia ed export", "CFO")
    ws['A4'].comment = Comment("EBITDA = Earnings Before Interest, Taxes, Depreciation, and Amortization", "CFO")

    # Formattazione
    for col in ['B', 'C']:
        for row in range(2, 11):
            cell = ws[f'{col}{row}']
            if cell.value and not isinstance(cell.value, str):
                cell.number_format = '#,##0'

    for row in range(2, 11):
        cell = ws[f'D{row}']
        if cell.value:
            cell.number_format = '0.00%'

    # Sheet 2: KPI
    ws2 = wb.create_sheet("KPI")
    ws2['A1'] = 'Indicatore'
    ws2['B1'] = 'Valore'
    ws2['C1'] = 'Target'
    ws2['D1'] = 'Performance'

    kpi_data = [
        ['ROE (%)', 15.2, 14.0, 'Sopra target'],
        ['ROI (%)', 12.5, 12.0, 'Sopra target'],
        ['PFN/EBITDA', 2.3, 3.0, 'OK'],
        ['DSO (giorni)', 65, 60, 'Da migliorare'],
        ['DPO (giorni)', 45, 50, 'OK']
    ]

    for i, row in enumerate(kpi_data, start=2):
        for j, value in enumerate(row, start=1):
            ws2.cell(row=i, column=j, value=value)

    # Salva file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    temp_file.close()

    return Path(temp_file.name)


def demo_excel_parser():
    """Dimostra l'uso del parser Excel con metadati completi."""
    print("\n" + "="*60)
    print("DEMO: Excel Parser con Metadati Completi")
    print("="*60)

    # Crea file di esempio
    excel_file = create_sample_excel()
    print(f"\n✅ File Excel creato: {excel_file}")

    # Inizializza parser
    parser = ExcelParser()

    # Estrai dati con metadati
    extracted_data, source_refs = parser.extract_with_source_references(excel_file)

    print("\n📊 Metadati Workbook:")
    print(f"  - File: {extracted_data.workbook_metadata.file_path}")
    print(f"  - Hash: {extracted_data.workbook_metadata.file_hash[:16]}...")
    print(f"  - Dimensione: {extracted_data.workbook_metadata.file_size} bytes")
    print(f"  - Fogli: {extracted_data.workbook_metadata.sheets_count}")

    print("\n📋 Fogli trovati:")
    for sheet in extracted_data.workbook_metadata.sheets:
        print(f"  - {sheet.name}:")
        print(f"    • Righe: {sheet.max_row}")
        print(f"    • Colonne: {sheet.max_column}")
        print(f"    • Formule: {sheet.formulas_count}")
        print(f"    • Commenti: {sheet.comments_count}")

    print(f"\n📑 Tabelle rilevate: {len(extracted_data.tables)}")
    for table in extracted_data.tables:
        print(f"  - {table.sheet_name}: {table.start_cell}:{table.end_cell}")
        print(f"    • Headers: {', '.join(table.headers[:3])}...")
        print(f"    • Righe: {table.total_rows}")

    print("\n💬 Commenti trovati:")
    for sheet_name, comments in extracted_data.comments.items():
        for cell_ref, comment in comments[:3]:
            print(f"  - {sheet_name}!{cell_ref}: '{comment[:50]}...'")

    print("\n📐 Formule trovate:")
    for sheet_name, formulas in extracted_data.formulas.items():
        print(f"  - {sheet_name}: {len(formulas)} formule")
        for cell_ref, formula in formulas[:3]:
            print(f"    • {cell_ref}: {formula}")

    print(f"\n🔗 Source References generate: {len(source_refs)}")
    for ref in source_refs[:3]:
        print(f"  - {ref.section} | {ref.subsection}")

    # Cleanup
    excel_file.unlink()

    return extracted_data


def demo_raw_blocks_extractor():
    """Dimostra l'uso del Raw Blocks Extractor."""
    print("\n" + "="*60)
    print("DEMO: Raw Blocks Extractor")
    print("="*60)

    # Crea file di esempio
    excel_file = create_sample_excel()
    print(f"\n✅ File creato: {excel_file}")

    # Inizializza extractor
    extractor = RawBlocksExtractor()

    # Estrai blocchi
    document_blocks = extractor.extract(excel_file)

    print("\n📄 Documento analizzato:")
    print(f"  - ID: {document_blocks.document_id}")
    print(f"  - Tipo: {document_blocks.document_type.value}")
    print(f"  - Pagine/Fogli: {len(document_blocks.pages)}")

    all_blocks = document_blocks.get_all_blocks()
    print(f"\n📦 Blocchi totali estratti: {len(all_blocks)}")

    # Analisi per tipo
    text_blocks = document_blocks.get_blocks_by_type(BlockType.TEXT)
    table_blocks = document_blocks.get_blocks_by_type(BlockType.TABLE)
    formula_blocks = document_blocks.get_blocks_by_type(BlockType.FORMULA)

    print("\n📊 Distribuzione blocchi per tipo:")
    print(f"  - Testo: {len(text_blocks)}")
    print(f"  - Tabelle: {len(table_blocks)}")
    print(f"  - Formule: {len(formula_blocks)}")

    print("\n🔍 Dettaglio blocchi per pagina/foglio:")
    for page in document_blocks.pages:
        page_name = page.sheet_name or f"Pagina {page.page_number}"
        print(f"\n  {page_name}:")
        print(f"    • Blocchi totali: {len(page.blocks)}")

        for block in page.blocks[:2]:  # Mostra primi 2 blocchi
            print(f"    • [{block.block_type.value}] {block.block_id}")
            if block.text_content:
                preview = block.text_content[:50] + "..." if len(block.text_content) > 50 else block.text_content
                print(f"      Contenuto: {preview}")
            print(f"      Source ref: {block.to_source_ref()}")

    # Salva blocchi in JSON
    blocks_file = Path("data/raw_blocks") / f"{excel_file.stem}_blocks.json"
    blocks_file.parent.mkdir(parents=True, exist_ok=True)
    document_blocks.save_to_json(blocks_file)
    print(f"\n💾 Blocchi salvati in: {blocks_file}")

    # Cleanup
    excel_file.unlink()

    return document_blocks


async def demo_enterprise_integration():
    """Dimostra l'integrazione con Enterprise Orchestrator."""
    print("\n" + "="*60)
    print("DEMO: Integrazione con Enterprise Orchestrator")
    print("="*60)

    # Crea file di esempio
    excel_file = create_sample_excel()
    print(f"\n✅ File Excel creato: {excel_file}")

    # Inizializza orchestrator
    orchestrator = EnterpriseOrchestrator()

    # Processa file Excel
    print("\n🔄 Processando file Excel con Enterprise Orchestrator...")
    extracted_data, source_refs = await orchestrator.process_excel_file(excel_file)

    print("\n✅ Elaborazione completata:")
    print(f"  - Source references: {len(source_refs)}")
    print(f"  - Fogli processati: {len(extracted_data.data_frames)}")

    # Estrai raw blocks
    print("\n🔄 Estraendo raw blocks...")
    blocks = await orchestrator.extract_raw_blocks(excel_file)

    print("\n✅ Estrazione completata:")
    print(f"  - Blocchi totali: {len(blocks.get_all_blocks())}")
    print(f"  - Blocchi testo: {len(blocks.get_blocks_by_type(BlockType.TEXT))}")
    print(f"  - Blocchi tabella: {len(blocks.get_blocks_by_type(BlockType.TABLE))}")

    # Mostra statistiche
    stats = orchestrator.get_processing_stats()
    print("\n📈 Statistiche di elaborazione:")
    for key, value in stats.items():
        if not isinstance(value, dict):
            print(f"  - {key}: {value}")

    # Cleanup
    excel_file.unlink()

    print("\n✅ Demo completata con successo!")


def main():
    """Esegue tutte le demo."""
    print("\n" + "🚀 "*20)
    print("DEMO DEI NUOVI COMPONENTI IMPLEMENTATI")
    print("Excel Parser & Raw Blocks Extractor")
    print("🚀 "*20)

    try:
        # Demo 1: Excel Parser
        demo_excel_parser()

        # Demo 2: Raw Blocks Extractor
        demo_raw_blocks_extractor()

        # Demo 3: Enterprise Integration
        asyncio.run(demo_enterprise_integration())

        print("\n" + "="*60)
        print("✅ TUTTE LE DEMO COMPLETATE CON SUCCESSO!")
        print("="*60)

        print("\n📌 Riepilogo funzionalità implementate:")
        print("  1. ✅ Excel Parser con metadati completi")
        print("     - Estrazione tabelle, formule, commenti")
        print("     - Preservazione metadati sheet/celle")
        print("     - Generazione source references")

        print("\n  2. ✅ Raw Blocks Extractor")
        print("     - Estrazione blocchi per pagina/foglio")
        print("     - Supporto PDF, Excel, HTML, testo")
        print("     - Serializzazione/deserializzazione JSON")

        print("\n  3. ✅ Integrazione Enterprise Orchestrator")
        print("     - Metodi process_excel_file() e extract_raw_blocks()")
        print("     - Salvataggio automatico blocchi")
        print("     - Statistiche di elaborazione")

    except Exception as e:
        print(f"\n❌ Errore durante la demo: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
