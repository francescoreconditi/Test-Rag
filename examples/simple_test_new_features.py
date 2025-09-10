# ============================================
# FILE DI TEST/DEBUG - NON PER PRODUZIONE
# Creato da: Claude Code
# Data: 2025-09-08
# Scopo: Test semplificato dei nuovi componenti
# ============================================

"""Test semplificato per Excel Parser e Raw Blocks Extractor."""

from pathlib import Path
import tempfile

from openpyxl import Workbook
from openpyxl.comments import Comment

# Import solo dei nuovi componenti
from src.application.parsers.excel_parser import ExcelParser
from src.application.services.raw_blocks_extractor import RawBlocksExtractor


def create_test_excel():
    """Crea un file Excel di test."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"

    # Dati semplici
    ws['A1'] = 'Metrica'
    ws['B1'] = 'Valore'
    ws['A2'] = 'Ricavi'
    ws['B2'] = 1000000
    ws['A3'] = 'Costi'
    ws['B3'] = 800000
    ws['A4'] = 'EBITDA'
    ws['B4'] = '=B2-B3'

    # Commento
    ws['A2'].comment = Comment("Test comment", "Author")

    # Salva
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)


def test_excel_parser():
    """Test Excel Parser."""
    print("\n=== TEST EXCEL PARSER ===")

    # Crea file
    excel_file = create_test_excel()
    print(f"OK File creato: {excel_file}")

    # Parse
    parser = ExcelParser()
    data, refs = parser.extract_with_source_references(excel_file)

    print("OK Parsing completato:")
    print(f"  - Fogli: {data.workbook_metadata.sheets_count}")
    print(f"  - Tabelle: {len(data.tables)}")
    print(f"  - Formule: {len(data.formulas.get('Test', []))}")
    print(f"  - Commenti: {len(data.comments.get('Test', []))}")
    print(f"  - Source refs: {len(refs)}")

    # Mostra alcuni dettagli
    if data.tables:
        table = data.tables[0]
        print("\n>> Prima tabella:")
        print(f"  - Range: {table.start_cell}:{table.end_cell}")
        print(f"  - Headers: {table.headers}")

    if 'Test' in data.formulas:
        print("\n>> Formule trovate:")
        for cell, formula in data.formulas['Test']:
            print(f"  - {cell}: {formula}")

    # Cleanup
    excel_file.unlink()
    return True


def test_raw_blocks():
    """Test Raw Blocks Extractor."""
    print("\n=== TEST RAW BLOCKS EXTRACTOR ===")

    # Crea file
    excel_file = create_test_excel()
    print(f"OK File creato: {excel_file}")

    # Estrai blocchi
    extractor = RawBlocksExtractor()
    blocks = extractor.extract(excel_file)

    print("OK Estrazione completata:")
    print(f"  - Documento ID: {blocks.document_id}")
    print(f"  - Tipo: {blocks.document_type.value}")
    print(f"  - Pagine: {len(blocks.pages)}")

    all_blocks = blocks.get_all_blocks()
    print(f"  - Blocchi totali: {len(all_blocks)}")

    # Mostra dettagli blocchi
    for block in all_blocks:
        print(f"\n>> Blocco {block.block_id}:")
        print(f"  - Tipo: {block.block_type.value}")
        print(f"  - Sheet: {block.sheet_name}")
        if block.metadata:
            print(f"  - Metadata: {block.metadata}")

    # Test serializzazione
    json_file = Path("test_blocks.json")
    blocks.save_to_json(json_file)
    print(f"\n>> Salvato in: {json_file}")

    # Test deserializzazione
    loaded = blocks.load_from_json(json_file)
    print(f"OK Ricaricato: {len(loaded.get_all_blocks())} blocchi")

    # Cleanup
    excel_file.unlink()
    json_file.unlink()
    return True


def main():
    """Esegue i test."""
    print("\n>>> TEST DEI NUOVI COMPONENTI <<<")
    print("="*40)

    try:
        # Test 1: Excel Parser
        if test_excel_parser():
            print("\n=> Excel Parser: PASSED")

        # Test 2: Raw Blocks Extractor
        if test_raw_blocks():
            print("\n=> Raw Blocks Extractor: PASSED")

        print("\n" + "="*40)
        print(">> TUTTI I TEST PASSATI! <<")
        print("\n=> Funzionalita' implementate:")
        print("  1. Parser Excel con metadati completi OK")
        print("  2. Estrazione raw blocks per pagina OK")

    except Exception as e:
        print(f"\n>> Errore: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
