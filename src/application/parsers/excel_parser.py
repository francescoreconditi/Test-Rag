# ============================================
# FILE DI TEST/DEBUG - NON PER PRODUZIONE
# Creato da: Claude Code
# Data: 2025-09-08
# Scopo: Parser Excel dedicato con preservazione completa dei metadati
# ============================================

"""
Excel Parser con preservazione completa dei metadati.

Estrae dati da file Excel preservando:
- Nome sheet e indice
- Range celle (A1:B10)
- Headers e loro posizione
- Formule e commenti
- Formattazione (valuta, percentuale, date)
- Celle unite (merged cells)
- Metadati del workbook
"""

import hashlib
import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.domain.value_objects.source_reference import SourceReference

logger = logging.getLogger(__name__)


@dataclass
class CellMetadata:
    """Metadati completi per una singola cella."""
    
    sheet_name: str
    cell_reference: str  # Es: "A1", "B12"
    row: int
    column: int
    value: Any
    formula: Optional[str] = None
    comment: Optional[str] = None
    number_format: Optional[str] = None
    data_type: Optional[str] = None
    is_merged: bool = False
    merge_range: Optional[str] = None
    hyperlink: Optional[str] = None
    font_bold: bool = False
    fill_color: Optional[str] = None
    
    def to_source_ref(self, file_path: str) -> str:
        """Genera source reference per questa cella."""
        return f"{file_path}|sheet:{self.sheet_name}|cell:{self.cell_reference}"


@dataclass
class TableMetadata:
    """Metadati per una tabella rilevata nel foglio."""
    
    sheet_name: str
    table_name: Optional[str]
    start_cell: str  # Es: "A1"
    end_cell: str    # Es: "F20"
    headers: List[str]
    header_row: int
    data_rows: Tuple[int, int]  # (start, end)
    total_rows: int
    total_columns: int
    has_totals: bool = False
    totals_row: Optional[int] = None
    
    def to_source_ref(self, file_path: str) -> str:
        """Genera source reference per questa tabella."""
        return f"{file_path}|sheet:{self.sheet_name}|range:{self.start_cell}:{self.end_cell}"


@dataclass
class SheetMetadata:
    """Metadati completi per un singolo foglio."""
    
    name: str
    index: int
    visible: bool
    protection: bool
    used_range: str  # Es: "A1:Z100"
    max_row: int
    max_column: int
    tables: List[TableMetadata] = field(default_factory=list)
    named_ranges: Dict[str, str] = field(default_factory=dict)
    charts_count: int = 0
    pivot_tables_count: int = 0
    formulas_count: int = 0
    comments_count: int = 0
    merged_cells_ranges: List[str] = field(default_factory=list)


@dataclass
class WorkbookMetadata:
    """Metadati del workbook Excel."""
    
    file_path: str
    file_hash: str
    file_size: int
    created: Optional[datetime]
    modified: Optional[datetime]
    author: Optional[str]
    last_modified_by: Optional[str]
    sheets_count: int
    sheets: List[SheetMetadata]
    named_ranges: Dict[str, str]
    has_macros: bool
    has_external_links: bool


@dataclass
class ExtractedData:
    """Dati estratti con metadati completi."""
    
    workbook_metadata: WorkbookMetadata
    data_frames: Dict[str, pd.DataFrame]  # sheet_name -> DataFrame
    cell_metadata: Dict[str, List[CellMetadata]]  # sheet_name -> cells
    tables: List[TableMetadata]
    formulas: Dict[str, List[Tuple[str, str]]]  # sheet -> [(cell, formula)]
    comments: Dict[str, List[Tuple[str, str]]]  # sheet -> [(cell, comment)]
    raw_values: Dict[str, Dict[str, Any]]  # sheet -> {cell_ref: value}


class ExcelParser:
    """Parser Excel avanzato con preservazione completa dei metadati."""
    
    def __init__(self):
        self.workbook = None
        self.file_path = None
        
    def parse(self, file_path: Union[str, Path]) -> ExtractedData:
        """
        Parsing completo del file Excel con estrazione di tutti i metadati.
        
        Args:
            file_path: Percorso del file Excel
            
        Returns:
            ExtractedData con tutti i dati e metadati estratti
        """
        self.file_path = Path(file_path)
        
        if not self.file_path.exists():
            raise FileNotFoundError(f"File Excel non trovato: {file_path}")
            
        logger.info(f"Parsing Excel file: {self.file_path}")
        
        # Carica workbook con openpyxl per metadati completi
        self.workbook = load_workbook(
            str(self.file_path), 
            data_only=False,  # Preserva formule
            keep_vba=True,    # Preserva macro
            keep_links=True   # Preserva link esterni
        )
        
        # Estrai metadati workbook
        workbook_metadata = self._extract_workbook_metadata()
        
        # Estrai dati e metadati per ogni sheet
        data_frames = {}
        cell_metadata = {}
        all_tables = []
        all_formulas = {}
        all_comments = {}
        raw_values = {}
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # DataFrame per analisi veloce
            df = pd.read_excel(
                str(self.file_path), 
                sheet_name=sheet_name,
                header=None,  # Non assumere headers
                keep_default_na=False
            )
            data_frames[sheet_name] = df
            
            # Metadati celle
            cells = self._extract_cell_metadata(sheet)
            cell_metadata[sheet_name] = cells
            
            # Rileva tabelle
            tables = self._detect_tables(sheet, df)
            all_tables.extend(tables)
            
            # Estrai formule
            formulas = self._extract_formulas(sheet)
            if formulas:
                all_formulas[sheet_name] = formulas
                
            # Estrai commenti
            comments = self._extract_comments(sheet)
            if comments:
                all_comments[sheet_name] = comments
                
            # Valori raw per riferimento
            raw_values[sheet_name] = self._extract_raw_values(sheet)
            
        return ExtractedData(
            workbook_metadata=workbook_metadata,
            data_frames=data_frames,
            cell_metadata=cell_metadata,
            tables=all_tables,
            formulas=all_formulas,
            comments=all_comments,
            raw_values=raw_values
        )
    
    def _extract_workbook_metadata(self) -> WorkbookMetadata:
        """Estrae metadati del workbook."""
        file_stats = self.file_path.stat()
        
        # Calcola hash del file
        with open(self.file_path, 'rb') as f:
            file_hash = hashlib.sha256(f.read()).hexdigest()
            
        # Metadati delle proprietà del documento
        props = self.workbook.properties
        
        # Metadati sheets
        sheets_metadata = []
        for idx, sheet_name in enumerate(self.workbook.sheetnames):
            sheet = self.workbook[sheet_name]
            sheet_meta = self._extract_sheet_metadata(sheet, idx)
            sheets_metadata.append(sheet_meta)
            
        return WorkbookMetadata(
            file_path=str(self.file_path),
            file_hash=file_hash,
            file_size=file_stats.st_size,
            created=props.created,
            modified=props.modified,
            author=props.creator,
            last_modified_by=props.lastModifiedBy,
            sheets_count=len(self.workbook.sheetnames),
            sheets=sheets_metadata,
            named_ranges={nr: str(self.workbook.defined_names[nr]) for nr in self.workbook.defined_names} if hasattr(self.workbook, 'defined_names') else {},
            has_macros=self.workbook.vba_archive is not None if hasattr(self.workbook, 'vba_archive') else False,
            has_external_links=bool(self.workbook.external_links) if hasattr(self.workbook, 'external_links') else False
        )
    
    def _extract_sheet_metadata(self, sheet: Worksheet, index: int) -> SheetMetadata:
        """Estrae metadati di un singolo sheet."""
        # Range utilizzato
        min_row = sheet.min_row or 1
        min_col = sheet.min_column or 1
        max_row = sheet.max_row or 1
        max_col = sheet.max_column or 1
        
        used_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        
        # Conta formule e commenti
        formulas_count = sum(1 for row in sheet.iter_rows() for cell in row if cell.value and str(cell.value).startswith('='))
        comments_count = sum(1 for row in sheet.iter_rows() for cell in row if cell.comment)
        
        # Merged cells
        merged_ranges = [str(range_) for range_ in sheet.merged_cells.ranges]
        
        return SheetMetadata(
            name=sheet.title,
            index=index,
            visible=sheet.sheet_state == 'visible',
            protection=sheet.protection.sheet,
            used_range=used_range,
            max_row=max_row,
            max_column=max_col,
            tables=[],  # Popolato dopo
            named_ranges={},  # Popolato dopo
            charts_count=len(sheet._charts) if hasattr(sheet, '_charts') else 0,
            pivot_tables_count=len(sheet._pivots) if hasattr(sheet, '_pivots') else 0,
            formulas_count=formulas_count,
            comments_count=comments_count,
            merged_cells_ranges=merged_ranges
        )
    
    def _extract_cell_metadata(self, sheet: Worksheet) -> List[CellMetadata]:
        """Estrae metadati dettagliati per ogni cella con valore."""
        cells_metadata = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Verifica se la cella è parte di un merge
                    is_merged = False
                    merge_range = None
                    for merged_range_obj in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range_obj:
                            is_merged = True
                            merge_range = str(merged_range_obj)
                            break
                    
                    metadata = CellMetadata(
                        sheet_name=sheet.title,
                        cell_reference=cell.coordinate,
                        row=cell.row,
                        column=cell.column,
                        value=cell.value,
                        formula=cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None,
                        comment=cell.comment.text if cell.comment else None,
                        number_format=cell.number_format,
                        data_type=cell.data_type,
                        is_merged=is_merged,
                        merge_range=merge_range,
                        hyperlink=cell.hyperlink.target if cell.hyperlink else None,
                        font_bold=cell.font.bold if cell.font else False,
                        fill_color=cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
                    )
                    cells_metadata.append(metadata)
                    
        return cells_metadata
    
    def _detect_tables(self, sheet: Worksheet, df: pd.DataFrame) -> List[TableMetadata]:
        """Rileva automaticamente le tabelle nel foglio."""
        tables = []
        
        # Strategia 1: Usa tabelle Excel definite
        for table in sheet.tables.values():
            table_range = table.ref
            start_cell = table_range.split(':')[0]
            end_cell = table_range.split(':')[1] if ':' in table_range else start_cell
            
            # Estrai headers
            header_row = sheet[table.headerRowCount or 1]
            headers = [cell.value for cell in header_row if cell.value]
            
            tables.append(TableMetadata(
                sheet_name=sheet.title,
                table_name=table.name,
                start_cell=start_cell,
                end_cell=end_cell,
                headers=headers,
                header_row=table.headerRowCount or 1,
                data_rows=(table.headerRowCount + 1, sheet.max_row),
                total_rows=sheet.max_row - table.headerRowCount,
                total_columns=len(headers),
                has_totals=table.totalsRowCount > 0 if table.totalsRowCount else False,
                totals_row=sheet.max_row if table.totalsRowCount else None
            ))
        
        # Strategia 2: Rileva tabelle basandosi su pattern (se non ci sono tabelle definite)
        if not tables:
            tables.extend(self._detect_tables_by_pattern(sheet, df))
            
        return tables
    
    def _detect_tables_by_pattern(self, sheet: Worksheet, df: pd.DataFrame) -> List[TableMetadata]:
        """Rileva tabelle basandosi su pattern di dati."""
        tables = []
        
        # Cerca righe che sembrano headers (testo in tutte le colonne consecutive)
        for row_idx in range(min(10, len(df))):  # Cerca nei primi 10 righe
            row = df.iloc[row_idx]
            
            # Conta celle non vuote consecutive
            non_empty = 0
            start_col = None
            
            for col_idx, value in enumerate(row):
                if pd.notna(value) and str(value).strip():
                    if start_col is None:
                        start_col = col_idx
                    non_empty += 1
                else:
                    if non_empty >= 3:  # Almeno 3 colonne per considerarla una tabella
                        # Verifica che ci siano dati sotto
                        if row_idx + 1 < len(df):
                            next_row = df.iloc[row_idx + 1]
                            if any(pd.notna(next_row[start_col:start_col + non_empty])):
                                # Trova fine tabella
                                end_row = row_idx + 1
                                for check_row in range(row_idx + 1, len(df)):
                                    if all(pd.isna(df.iloc[check_row][start_col:start_col + non_empty])):
                                        break
                                    end_row = check_row
                                
                                headers = [str(v) for v in row[start_col:start_col + non_empty] if pd.notna(v)]
                                
                                tables.append(TableMetadata(
                                    sheet_name=sheet.title,
                                    table_name=None,
                                    start_cell=f"{get_column_letter(start_col + 1)}{row_idx + 1}",
                                    end_cell=f"{get_column_letter(start_col + non_empty)}{end_row + 1}",
                                    headers=headers,
                                    header_row=row_idx + 1,
                                    data_rows=(row_idx + 2, end_row + 1),
                                    total_rows=end_row - row_idx,
                                    total_columns=non_empty,
                                    has_totals=self._check_for_totals(df, end_row, start_col, non_empty),
                                    totals_row=end_row + 1 if self._check_for_totals(df, end_row, start_col, non_empty) else None
                                ))
                    
                    non_empty = 0
                    start_col = None
                    
        return tables
    
    def _check_for_totals(self, df: pd.DataFrame, last_row: int, start_col: int, num_cols: int) -> bool:
        """Verifica se l'ultima riga contiene totali."""
        if last_row >= len(df):
            return False
            
        last_row_data = df.iloc[last_row][start_col:start_col + num_cols]
        
        # Cerca parole chiave per totali
        totals_keywords = ['total', 'totale', 'sum', 'somma', 'subtotal', 'subtotale']
        
        for value in last_row_data:
            if pd.notna(value) and any(keyword in str(value).lower() for keyword in totals_keywords):
                return True
                
        return False
    
    def _extract_formulas(self, sheet: Worksheet) -> List[Tuple[str, str]]:
        """Estrae tutte le formule dal foglio."""
        formulas = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.append((cell.coordinate, cell.value))
                    
        return formulas
    
    def _extract_comments(self, sheet: Worksheet) -> List[Tuple[str, str]]:
        """Estrae tutti i commenti dal foglio."""
        comments = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.comment:
                    comments.append((cell.coordinate, cell.comment.text))
                    
        return comments
    
    def _extract_raw_values(self, sheet: Worksheet) -> Dict[str, Any]:
        """Estrae valori raw per riferimento veloce."""
        raw_values = {}
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    raw_values[cell.coordinate] = cell.value
                    
        return raw_values
    
    def extract_with_source_references(self, file_path: Union[str, Path]) -> Tuple[ExtractedData, List[SourceReference]]:
        """
        Estrae dati con generazione automatica di SourceReference.
        
        Returns:
            Tuple di ExtractedData e lista di SourceReference
        """
        data = self.parse(file_path)
        source_refs = []
        
        # Genera source references per ogni tabella
        for table in data.tables:
            source_ref = SourceReference(
                file_path=data.workbook_metadata.file_path,
                page_number=None,  # Excel non ha pagine
                sheet=table.sheet_name,
                extraction_method=f"Table: {table.table_name}" if table.table_name else f"Range: {table.start_cell}:{table.end_cell}"
            )
            source_refs.append(source_ref)
            
        # Genera source references per celle importanti (con formule o commenti)
        for sheet_name, cells in data.cell_metadata.items():
            for cell in cells:
                if cell.formula or cell.comment:
                    source_ref = SourceReference(
                        file_path=data.workbook_metadata.file_path,
                        page_number=None,
                        sheet=sheet_name,
                        cell=cell.cell_reference,
                        extraction_method=f"Cell with {'formula' if cell.formula else 'comment'}"
                    )
                    source_refs.append(source_ref)
                    
        return data, source_refs